import csv
import io
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook
from email_sender.models import Contact
from django.contrib.auth.models import User
from email_sender.models import ContactGroup


def import_contacts(file, owner:User, group:ContactGroup|None = None)->dict:
    """
    импорт контактов из CSV или XLSX файла.
    необходимые колонки: name, email, description
    """
    filename = file.name.lower()
    if filename.endswith(".csv"):
        rows = read_csv(file)
    elif filename.endswith(".xlsx"):
        rows = read_xlsx(file)
    else:
        raise ValueError("только CSV и XLSX файлы")

    return create_contacts(
        rows=rows,
        owner=owner,
        group=group,
    )


def read_csv(file)-> list:
    """
    читает CSV и возвращает список строк
    """
    file.seek(0)
    content = file.read().decode("utf-8-sig")
    text_file = io.StringIO(content)
    reader = csv.DictReader(text_file)
    validate_headers(reader.fieldnames)
    return list(reader)


def read_xlsx(file)-> list:
    """
    читает XLSX и возвращает список строк
    """
    file.seek(0)
    workbook = load_workbook(
        filename=file,
        read_only=True,
        data_only=True,
    )
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError("файл пуст")

    headers = [
        str(header).strip().lower()
        if header is not None
        else ""
        for header in headers
    ]
    validate_headers(headers)
    result = []
    for row in rows:
        row_data = dict(zip(headers, row))
        result.append(row_data)
    workbook.close()
    return result


def validate_headers(headers:list|None) -> None:
    """
    проверка наличия обязательных колонок
    """
    if not headers:
        raise ValueError("Файл не содержит заголовков")

    normalized_headers = [
        str(header).strip().lower()
        for header in headers
        if header
    ]
    required_headers = {
        "name",
        "email",
    }
    missing_headers = (required_headers - set(normalized_headers))
    if missing_headers:
        missing = ", ".join(sorted(missing_headers))
        raise ValueError(f"отсутствуют обязательные колонки: {missing}")


def create_contacts(rows:list,owner:User,group:ContactGroup|None = None)->dict:
    """
    создание контакта из полученных строк
    """
    total_count = 0
    created_count = 0
    skipped_count = 0
    error_count = 0

    errors = []

    # данные начинаются со второй строки тк на первой заголовки
    for row_number, row in enumerate(rows,start=2):
        total_count += 1
        name = get_cell_value(row.get("name"))
        email = get_cell_value(row.get("email")).lower()
        description = get_cell_value(row.get("description"))
        if not name:
            error_count += 1
            errors.append(
                {
                    "row": row_number,
                    "error": "имя контакта отсутствует",
                }
            )
            continue

        if not email:
            error_count += 1
            errors.append(
                {
                    "row": row_number,
                    "error": "Email отсутствует.",
                }
            )
            continue

        try:
            validate_email(email)
        except ValidationError:
            error_count += 1
            errors.append(
                {
                    "row": row_number,
                    "email": email,
                    "error": "Некорректный email.",
                }
            )
            continue

        contact_exists = Contact.objects.filter(owner=owner, email__iexact=email).exists()

        if contact_exists:
            skipped_count += 1
            continue

        contact = Contact.objects.create(
            owner=owner,
            name=name,
            email=email,
            description=description,
        )

        if group is not None:
            group.contacts.add(contact)
        created_count += 1

    return {
        "total": total_count,
        "created": created_count,
        "skipped": skipped_count,
        "errors_count": error_count,
        "errors": errors,
    }



def get_cell_value(value)->str:
    """
    преобразование данных из ячейки в строку
    """
    if value is None:
        return ""

    return str(value).strip()